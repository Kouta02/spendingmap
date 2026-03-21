"""
Management command para importar categorias e bancos da planilha .xlsx.
Importa apenas dados únicos, sem duplicar registros existentes.

Uso:
  python manage.py import_from_xlsx docs/despesas_mensais_exemplo.xlsx
"""
from django.core.management.base import BaseCommand

import openpyxl

from apps.categories.models import Category
from apps.banks.models import Bank


# Mapeamento de cores padrão para categorias
CATEGORY_COLORS = {
    'Alimentação': '#ff9800',
    'Assinatura': '#9c27b0',
    'Combustível': '#795548',
    'Contas básicas': '#f44336',
    'Educação': '#2196f3',
    'Eletrônicos': '#607d8b',
    'Estacionamento': '#9e9e9e',
    'Lazer': '#4caf50',
    'Mercado/Padaria': '#ff5722',
    'Obras': '#8d6e63',
    'Outros': '#bdbdbd',
    'Presentes': '#e91e63',
    'Saúde': '#00bcd4',
    'Tarifas': '#ff7043',
    'Transporte': '#3f51b5',
    'Vestuários e similares': '#673ab7',
    'Veículos': '#455a64',
    'Viagem': '#009688',
}

CATEGORY_ICONS = {
    'Alimentação': 'restaurant',
    'Assinatura': 'subscriptions',
    'Combustível': 'local_gas_station',
    'Contas básicas': 'receipt',
    'Educação': 'school',
    'Eletrônicos': 'devices',
    'Estacionamento': 'local_parking',
    'Lazer': 'sports_esports',
    'Mercado/Padaria': 'shopping_cart',
    'Obras': 'construction',
    'Outros': 'more_horiz',
    'Presentes': 'card_giftcard',
    'Saúde': 'health_and_safety',
    'Tarifas': 'account_balance',
    'Transporte': 'directions_car',
    'Vestuários e similares': 'checkroom',
    'Veículos': 'two_wheeler',
    'Viagem': 'flight',
}

BANK_COLORS = {
    'Banco do Brasil': '#ffe100',
    'Bradesco': '#cc092f',
    'NuBank': '#820ad1',
}


class Command(BaseCommand):
    help = 'Importa categorias e bancos da planilha .xlsx'

    def add_arguments(self, parser):
        parser.add_argument('filepath', type=str, help='Caminho para o arquivo .xlsx')

    def handle(self, *args, **options):
        filepath = options['filepath']

        try:
            wb = openpyxl.load_workbook(filepath)
        except FileNotFoundError:
            self.stderr.write(self.style.ERROR(f'Arquivo não encontrado: {filepath}'))
            return

        categories_found = set()
        banks_found = set()

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                cat = row[2]
                bank = row[4]
                if cat and isinstance(cat, str):
                    cleaned = cat.strip()
                    if not cleaned.startswith('='):
                        categories_found.add(cleaned)
                if bank and isinstance(bank, str):
                    banks_found.add(bank.strip())

        # Importar categorias
        cat_created = 0
        for cat_name in sorted(categories_found):
            _, created = Category.objects.get_or_create(
                name=cat_name,
                parent=None,
                defaults={
                    'color': CATEGORY_COLORS.get(cat_name, ''),
                    'icon': CATEGORY_ICONS.get(cat_name, ''),
                },
            )
            if created:
                cat_created += 1
                self.stdout.write(f'  + Categoria: {cat_name}')
            else:
                self.stdout.write(f'  = Categoria já existe: {cat_name}')

        # Importar bancos
        bank_created = 0
        for bank_name in sorted(banks_found):
            _, created = Bank.objects.get_or_create(
                name=bank_name,
                defaults={
                    'color': BANK_COLORS.get(bank_name, ''),
                },
            )
            if created:
                bank_created += 1
                self.stdout.write(f'  + Banco: {bank_name}')
            else:
                self.stdout.write(f'  = Banco já existe: {bank_name}')

        self.stdout.write(self.style.SUCCESS(
            f'\nImportação concluída: {cat_created} categorias e {bank_created} bancos criados.'
        ))
